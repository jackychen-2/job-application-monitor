"""Excel export for application data using openpyxl."""

from __future__ import annotations

import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from job_monitor.models import Application

_HEADERS = [
    "ID", "Company", "Job Title", "Status", "Email Subject",
    "Email Date", "Source", "Notes", "Created At", "Updated At",
]

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def export_applications_excel(applications: List[Application]) -> bytes:
    """Export a list of Application ORM objects to an Excel (.xlsx) byte buffer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Header row
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, app in enumerate(applications, start=2):
        ws.cell(row=row_idx, column=1, value=app.id)
        ws.cell(row=row_idx, column=2, value=app.company)
        ws.cell(row=row_idx, column=3, value=app.job_title or "")
        ws.cell(row=row_idx, column=4, value=app.status)
        ws.cell(row=row_idx, column=5, value=app.email_subject or "")
        ws.cell(row=row_idx, column=6, value=app.email_date.isoformat() if app.email_date else "")
        ws.cell(row=row_idx, column=7, value=app.source)
        ws.cell(row=row_idx, column=8, value=app.notes or "")
        ws.cell(row=row_idx, column=9, value=app.created_at.isoformat() if app.created_at else "")
        ws.cell(row=row_idx, column=10, value=app.updated_at.isoformat() if app.updated_at else "")

    # Auto-fit column widths (approximate)
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(header) + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
